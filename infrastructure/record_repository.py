"""
Capa de INFRAESTRUCTURA.

Persiste los registros de inspección (los que el usuario confirma con el
botón "Agregar registro" en la GUI) como filas de un reporte Excel (.xlsx)
local: un solo archivo que va creciendo, una fila por pieza inspeccionada.
"""
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook


@dataclass
class InspectionRecord:
    tooth_count: int
    gear_type: str
    diameter_mm: Optional[float]
    corrosion: str          # "Sí", "No" o "N/D"
    quality: str             # "Aprobado" o "Defectuoso"
    lote: str = ""
    timestamp: str = ""

    def __post_init__(self) -> None:
        if not self.timestamp:
            self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class RecordRepository:
    """Escribe InspectionRecord como filas de un .xlsx, creando el archivo
    y su encabezado la primera vez que se usa."""

    _HEADERS = [
        "Fecha/Hora", "Dientes", "Tipo de engrane", "Diametro (mm)",
        "Corrosion", "Calidad", "Lote",
    ]

    def __init__(self, xlsx_path: str = "reporte_inspecciones.xlsx"):
        self._xlsx_path = xlsx_path

    def add(self, record: InspectionRecord) -> None:
        if os.path.isfile(self._xlsx_path):
            workbook = load_workbook(self._xlsx_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Inspecciones"
            sheet.append(self._HEADERS)

        sheet.append([
            record.timestamp,
            record.tooth_count,
            record.gear_type,
            round(record.diameter_mm, 2) if record.diameter_mm is not None else "N/D",
            record.corrosion,
            record.quality,
            record.lote,
        ])
        workbook.save(self._xlsx_path)

    @property
    def xlsx_path(self) -> str:
        return self._xlsx_path
